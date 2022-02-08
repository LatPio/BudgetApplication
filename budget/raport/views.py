import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from expenses.models import Category, Expense, Budynek
from userincome.models import UserIncome, Source
from django.contrib import messages
from django.core.paginator import Paginator
import os
import json
from django.conf import settings
from django.http import JsonResponse
from django.http import HttpResponse
from userpreferences.models import UserPreferences
import csv
import xlwt
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from itertools import chain
from django.db.models.functions import ExtractYear
from django.db.models import Avg, Max, Min, Sum


@login_required(login_url='/authentication/login')
def index(request):
    allDate = chain((Expense.objects.filter(owner=request.user).annotate(year=ExtractYear('date'))),
                    (UserIncome.objects.filter(owner=request.user).annotate(year=ExtractYear('date'))))
    unique_years = set()

    year = request.GET.get('year')
    print(year)
    for x in allDate:
        unique_years.add(x.year)
    context = {
        'roki': unique_years
    }
    return render(request, 'raport/index.html', context)



@login_required(login_url='/authentication/login')
def export_excel(request):



    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Zestawienie Transakcji ' + str(datetime.datetime.now()) + '.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Wydatki')
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Kwota', 'Waluta', 'Lokalizajca', 'Kategoria', 'Opis', 'Data']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    rows = Expense.objects.filter(owner=request.user).values_list('amount', 'waluta', 'budynek', 'category', 'description', 'date')

    for row in rows:
        row_num+=1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)

    # wb.save(response)


    ws2 = wb.add_sheet('Przychody')
    row_num1 = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    #
    columns1 = ['Kwota', 'Waluta', 'Lokalizajca', 'Źródło', 'Opis', 'Data']
    #
    for col_num in range(len(columns1)):
        ws2.write(row_num1, col_num, columns1[col_num], font_style)

    font_style = xlwt.XFStyle()

    rows1 = UserIncome.objects.filter(owner=request.user).values_list('amount', 'waluta', 'budynek', 'source',
                                                                  'description', 'date')

    for row in rows1:
        row_num1 += 1

        for col_num in range(len(row)):
            ws2.write(row_num1, col_num, str(row[col_num]), font_style)
    #
    #
    wb.save(response)

    return response

@login_required(login_url='/authentication/login')
def excel_preaty_raport(request, rok):
    waluta_usera = UserPreferences.objects.get(user=request.user)
    # categories = Category.objects.filter(owner=request.user)
    # budynek = Budynek.objects.filter(owner=request.user)
    # sources = Source.objects.filter(owner=request.user)

    categories1 = Expense.objects.filter(owner=request.user).values_list('category', flat=True)
    categories = set()
    for x in categories1:
        categories.add(x)
    sources1 = UserIncome.objects.filter(owner=request.user).values_list('source', flat=True)
    sources = set()
    for x in sources1:
        sources.add(x)

    budynek1 = chain((Expense.objects.filter(owner=request.user).values_list('budynek', flat=True)),
                     (UserIncome.objects.filter(owner=request.user).values_list('budynek', flat=True)))
    budynek = set()

    for x in budynek1:
        budynek.add(x)



    year = rok

    print(year)

    namesOfLocations = budynek



    workbook = Workbook()
    worksheet1 = workbook.active
    worksheet1.title='Podsumowanie'

    for i in namesOfLocations:
        workbook.create_sheet(i)  # or template.create_sheet(title=worksheets[i])
    for location in namesOfLocations:

        information_data_income = {}

        for sour in sources:
            sourcesMonhlydata = []
            for miech in range(1, 13):
                information = (UserIncome.objects.filter(owner=request.user).filter(
                    date__year=year).filter(date__month=miech).filter(waluta=waluta_usera.currency)).filter(
                    source=sour).filter(budynek=location).aggregate(Sum('amount'))
                if information['amount__sum'] is None:
                    information['amount__sum'] = 0

                sourcesMonhlydata.append(float(information['amount__sum']))
            information_data_income[sour] = sourcesMonhlydata

        information_data_expanse = {}
        for cat in categories:
            categoryMonhlydata = []
            for miech in range(1, 13):
                information = (Expense.objects.filter(owner=request.user).filter(
                    date__year=year).filter(date__month=miech).filter(waluta=waluta_usera.currency)).filter(
                    category=cat).filter(budynek=location).aggregate(Sum('amount'))
                if information['amount__sum'] is None:
                    information['amount__sum'] = 0
                categoryMonhlydata.append(float(information['amount__sum']))
            information_data_expanse[cat] = categoryMonhlydata

        headings = ['Kategoria', 'Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec', 'Lipiec', 'Sierpierń',
                    'Wrzesień',
                    'Październik', 'Listopad', 'Grudzień']

        workbook[location].append([('Wpływy w ' + location + ' za rok: ' + year)])
        workbook[location].append(headings)
        for source in information_data_income:
            data = information_data_income[source]
            workbook[location].append([source] + data)

        workbook[location].append([('Wydatki w ' + location + ' za rok: ' + year)])
        workbook[location].append(headings)
        for category in information_data_expanse:
            data2 = information_data_expanse[category]
            workbook[location].append([category] + data2)

        datalenght = len(information_data_income) + len(information_data_expanse) + 5
        for row in range(1, datalenght):
            workbook[location]['a' + str(row)].font = Font(bold=True)

        workbook[location]['a1'].font = Font(size=20)
        position = len(information_data_income) + 3

        workbook[location]['a' + str(position)].font = Font(size=20)





    # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    information1 = {}
    for sour in sources:
        sourcesMonhlydata = []
        for miech in range(1, 13):
            information = (UserIncome.objects.filter(owner=request.user).filter(
                date__year=year).filter(date__month=miech).filter(waluta=waluta_usera.currency)).filter(
                source=sour).aggregate(Sum('amount'))
            if information['amount__sum'] is None:
                information['amount__sum'] = 0

            sourcesMonhlydata.append(float(information['amount__sum']))
        information1[sour] = sourcesMonhlydata

    information2 = {}
    for cat in categories:
        categoryMonhlydata = []
        for miech in range(1, 13):
            information = (Expense.objects.filter(owner=request.user).filter(
                date__year=year).filter(date__month=miech).filter(waluta=waluta_usera.currency)).filter(
                category=cat).aggregate(Sum('amount'))
            if information['amount__sum'] is None:
                information['amount__sum'] = 0
            categoryMonhlydata.append(float(information['amount__sum']))
        information2[cat] = categoryMonhlydata

    headings = ['Kategoria', 'Styczeń', 'Luty', 'Marzec', 'Kwiecień', 'Maj', 'Czerwiec', 'Lipiec', 'Sierpierń', 'Wrzesień',
              'Październik', 'Listopad', 'Grudzień']

    worksheet1.append([('Wpływy za rok: ' + year)])
    worksheet1.append(headings)
    for source in information1:
        data = information1[source]
        worksheet1.append([source] + data)

    worksheet1.append([('Wydatki za rok: ' + year)])
    worksheet1.append(headings)
    for category in information2:
        data2 = information2[category]
        worksheet1.append([category] + data2)

    datalenght = len(information1) + len(information2) + 5
    for row in range(1, datalenght):
        worksheet1['a' + str(row)].font = Font(bold=True)

    worksheet1['a1'].font = Font(size=20)
    position = len(information1) + 3

    worksheet1['a' + str(position)].font = Font(size=20)

    response = HttpResponse(content=save_virtual_workbook(workbook))
    response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'

    return response


